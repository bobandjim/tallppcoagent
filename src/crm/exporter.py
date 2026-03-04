"""
Excel exporter — writes validated records back to the 6-sheet import.xlsx format.

Usage:
    from src.crm.exporter import CRMExporter, ExportPayload

    payload = ExportPayload(venues=[...], contacts=[...], gig_leads=[...])
    CRMExporter("data/crm/leads.xlsx").write(payload)
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .schema import CallRecord, ContactRecord, GigLeadRecord, NoteRecord, VenueRecord


@dataclass
class ExportPayload:
    venues: list[VenueRecord] = field(default_factory=list)
    contacts: list[ContactRecord] = field(default_factory=list)
    gig_leads: list[GigLeadRecord] = field(default_factory=list)
    calls: list[CallRecord] = field(default_factory=list)
    notes: list[NoteRecord] = field(default_factory=list)


# Header row definitions (must match CRMImporter expectations)
COMPANIES_HEADERS = [
    "UniqueId", "Name", "Website", "Record Type", "Principal", "Phone", "Fax",
    "Rating", "Size", "Custom1", "Custom2", "Custom3", "Custom4", "Custom5",
    "Custom6", "Custom7", "Custom8", "Custom9", "Custom10",
    "Billing Street", "Billing City", "Billing State", "Billing Postcode",
    "Billing Country", "Notes",
]

CONTACTS_HEADERS = [
    "UniqueId", "First Name", "Last Name", "Company UniqueId", "Lead Source",
    "Work Phone", "Extension", "Mobile Phone", "Home Phone", "Other Phone", "Fax",
    "Email", "Phone Opt Out", "Email Opt Out",
    "Primary Street", "Primary City", "Primary State", "Primary Postcode", "Primary Country",
    "Notes", "Tag1", "Tag2", "Tag3", "Prefix", "Salutation", "Title", "Birthdate",
    "AltEmail",
    "Custom1", "Custom2", "Custom3", "Custom4", "Custom5",
    "Custom6", "Custom7", "Custom8", "Custom9", "Custom10",
]

GIGS_LEADS_HEADERS = [
    "UniqueId", "Name", "Status", "Company UniqueId", "Primary Contact UniqueId",
    "Secondary Contact UniqueId", "Lead Source", "GigID", "Sponsor",
    "Show Date", "Show Time", "Show Duration", "Event Type", "Guest Age Range",
    "Event Size", "Audience Type", "Guest of Honor", "Guest of Honor Age",
    "Upsell Quantity", "Alternate Showtimes", "Show Details",
    "Venue Address", "Site Contact", "Venue Name", "Travel Info",
    "Taxable Miles", "Billable Miles", "Milage Rate",
    "Fee", "Deposit", "Balance Paid", "Deposit Paid", "Gratuity Received",
    "Contract sent", "Contract Received", "Other Sales", "Coupon Code",
    "Next Step", "Decision Maker", "Decision Date", "Best time to contact",
    "Client Comment", "Quote1", "Description 1 ", "Quote2", "Description2",
    "Quote3", "Description3 ", "Upsell Price", "Upsell Description", "Deposit Percentage",
    "Custom1", "Custom2", "Custom3", "Custom4", "Custom5",
    "Custom6", "Custom7", "Custom8", "Custom9", "Custom10",
    "Custom11", "Custom12", "Custom13", "Custom14", "Custom15",
]

CALLS_HEADERS = [
    "UniqueId", "Subject", "CallDateTime", "Call Status", "Call Notes",
    "Related Gig", "Related Contact", "Related Company",
]

NOTES_HEADERS = [
    "UniqueId", "Subject", "Date", "Notes",
    "Related Gig", "Related Contact", "Related Company",
]

LOOKUPS_DATA = [
    ("YesNo", "Call", "Gig"),
    ("Yes", "Held", "Gig"),
    ("No", "Planned", "Lead"),
    (None, "Not Held", None),
]

HEADER_FILL = PatternFill("solid", fgColor="4F81BD")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _write_sheet(ws, headers: list[str], rows: list[list]):
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(row)
    for i, _ in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = 18


def _pad(row: list, length: int) -> list:
    """Pad row with None to match header length."""
    return row + [None] * max(0, length - len(row))


class CRMExporter:
    """Writes ExportPayload records into a properly structured import.xlsx workbook."""

    def __init__(self, path: str | Path):
        self.path = Path(path)
        self.path.parent.mkdir(parents=True, exist_ok=True)

    def write(self, payload: ExportPayload) -> None:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default sheet

        self._write_companies(wb.create_sheet("Companies"), payload.venues)
        self._write_contacts(wb.create_sheet("Contacts"), payload.contacts)
        self._write_gigs_leads(wb.create_sheet("Gigs_Leads"), payload.gig_leads)
        self._write_calls(wb.create_sheet("Calls"), payload.calls)
        self._write_notes(wb.create_sheet("Notes"), payload.notes)
        self._write_lookups(wb.create_sheet("Lookups"))

        wb.save(self.path)

    def _write_companies(self, ws, venues: list[VenueRecord]):
        rows = []
        for v in venues:
            row = [None] * len(COMPANIES_HEADERS)
            row[COMPANIES_HEADERS.index("UniqueId")] = v.lead_id
            row[COMPANIES_HEADERS.index("Name")] = v.company
            row[COMPANIES_HEADERS.index("Website")] = v.website
            row[COMPANIES_HEADERS.index("Record Type")] = v.segment
            row[COMPANIES_HEADERS.index("Phone")] = v.phone
            row[COMPANIES_HEADERS.index("Billing City")] = v.city
            row[COMPANIES_HEADERS.index("Billing State")] = v.state
            row[COMPANIES_HEADERS.index("Notes")] = v.notes
            row[COMPANIES_HEADERS.index("Custom1")] = v.seating_capacity_tier
            row[COMPANIES_HEADERS.index("Custom2")] = v.mission_excerpt
            row[COMPANIES_HEADERS.index("Custom3")] = v.programming_focus
            row[COMPANIES_HEADERS.index("Custom4")] = v.venue_name
            rows.append(row)
        _write_sheet(ws, COMPANIES_HEADERS, rows)

    def _write_contacts(self, ws, contacts: list[ContactRecord]):
        rows = []
        for c in contacts:
            row = [None] * len(CONTACTS_HEADERS)
            row[CONTACTS_HEADERS.index("UniqueId")] = c.contact_id
            row[CONTACTS_HEADERS.index("First Name")] = c.first_name
            row[CONTACTS_HEADERS.index("Last Name")] = c.last_name
            row[CONTACTS_HEADERS.index("Company UniqueId")] = c.company_id
            row[CONTACTS_HEADERS.index("Lead Source")] = c.source
            row[CONTACTS_HEADERS.index("Work Phone")] = c.phone
            row[CONTACTS_HEADERS.index("Email")] = c.email
            row[CONTACTS_HEADERS.index("Email Opt Out")] = c.email_opt_out
            row[CONTACTS_HEADERS.index("Title")] = c.title
            row[CONTACTS_HEADERS.index("Custom1")] = c.sequence_id
            row[CONTACTS_HEADERS.index("Custom2")] = c.step_number
            row[CONTACTS_HEADERS.index("Custom3")] = c.email_open
            row[CONTACTS_HEADERS.index("Custom4")] = c.email_click
            row[CONTACTS_HEADERS.index("Custom5")] = c.email_reply
            row[CONTACTS_HEADERS.index("Custom6")] = c.reply_category
            row[CONTACTS_HEADERS.index("Custom7")] = c.last_updated
            row[CONTACTS_HEADERS.index("Custom8")] = c.audit_trail_id
            rows.append(row)
        _write_sheet(ws, CONTACTS_HEADERS, rows)

    def _write_gigs_leads(self, ws, gig_leads: list[GigLeadRecord]):
        rows = []
        for g in gig_leads:
            row = [None] * len(GIGS_LEADS_HEADERS)
            row[GIGS_LEADS_HEADERS.index("UniqueId")] = g.gig_lead_id
            row[GIGS_LEADS_HEADERS.index("Name")] = g.name
            row[GIGS_LEADS_HEADERS.index("Status")] = g.status
            row[GIGS_LEADS_HEADERS.index("Company UniqueId")] = g.company_id
            row[GIGS_LEADS_HEADERS.index("Primary Contact UniqueId")] = g.contact_id
            row[GIGS_LEADS_HEADERS.index("Lead Source")] = g.source
            row[GIGS_LEADS_HEADERS.index("Show Date")] = g.show_date
            row[GIGS_LEADS_HEADERS.index("Venue Name")] = g.venue_name
            row[GIGS_LEADS_HEADERS.index("Fee")] = g.final_fee
            row[GIGS_LEADS_HEADERS.index("Quote1")] = g.projected_fee
            row[GIGS_LEADS_HEADERS.index("Deposit")] = g.deposit
            row[GIGS_LEADS_HEADERS.index("Contract sent")] = g.contract_sent
            row[GIGS_LEADS_HEADERS.index("Contract Received")] = g.contract_received
            row[GIGS_LEADS_HEADERS.index("Custom1")] = g.outcome
            row[GIGS_LEADS_HEADERS.index("Custom2")] = g.meeting_booked
            row[GIGS_LEADS_HEADERS.index("Custom3")] = g.show_booked
            row[GIGS_LEADS_HEADERS.index("Custom4")] = g.send_time
            rows.append(row)
        _write_sheet(ws, GIGS_LEADS_HEADERS, rows)

    def _write_calls(self, ws, calls: list[CallRecord]):
        rows = [
            [c.call_id, c.subject, c.call_datetime, c.status, c.notes,
             c.gig_id, c.contact_id, c.company_id]
            for c in calls
        ]
        _write_sheet(ws, CALLS_HEADERS, rows)

    def _write_notes(self, ws, notes: list[NoteRecord]):
        rows = [
            [n.note_id, n.subject, n.date, n.notes,
             n.gig_id, n.contact_id, n.company_id]
            for n in notes
        ]
        _write_sheet(ws, NOTES_HEADERS, rows)

    def _write_lookups(self, ws):
        _write_sheet(ws, ["YesNo", "Call", "Gig"], [list(r) for r in LOOKUPS_DATA])
