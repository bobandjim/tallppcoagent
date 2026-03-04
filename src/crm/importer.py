"""
Excel importer — reads the 6-sheet import.xlsx and returns validated Pydantic records.

Usage:
    from src.crm.importer import CRMImporter

    importer = CRMImporter("data/crm/leads.xlsx")
    result = importer.load_all()
    print(result.venues, result.contacts, result.gig_leads)
    print(result.errors)
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional
from datetime import datetime

import openpyxl
from pydantic import ValidationError

from .schema import CallRecord, ContactRecord, GigLeadRecord, NoteRecord, VenueRecord

logger = logging.getLogger(__name__)


@dataclass
class ImportResult:
    venues: list[VenueRecord] = field(default_factory=list)
    contacts: list[ContactRecord] = field(default_factory=list)
    gig_leads: list[GigLeadRecord] = field(default_factory=list)
    calls: list[CallRecord] = field(default_factory=list)
    notes: list[NoteRecord] = field(default_factory=list)
    errors: list[dict] = field(default_factory=list)

    @property
    def error_count(self) -> int:
        return len(self.errors)

    @property
    def total_records(self) -> int:
        return (
            len(self.venues)
            + len(self.contacts)
            + len(self.gig_leads)
            + len(self.calls)
            + len(self.notes)
        )


def _sheet_to_dicts(ws) -> list[dict[str, Any]]:
    """Convert a worksheet to a list of dicts keyed by header row values."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else f"_col_{i}" for i, h in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        result.append(dict(zip(headers, row)))
    return result


def _safe(d: dict, key: str, default=None):
    """Get a value from dict, returning default for None or missing."""
    v = d.get(key)
    return default if v is None else v


def _str(d: dict, key: str) -> Optional[str]:
    v = _safe(d, key)
    return str(v).strip() if v is not None else None


def _int(d: dict, key: str) -> Optional[int]:
    v = _safe(d, key)
    if v is None:
        return None
    try:
        return int(v)
    except (ValueError, TypeError):
        return None


def _float(d: dict, key: str) -> Optional[float]:
    v = _safe(d, key)
    if v is None:
        return None
    try:
        return float(str(v).replace(",", "").replace("$", ""))
    except (ValueError, TypeError):
        return None


def _yes_no(d: dict, key: str, default: str = "No") -> str:
    v = _safe(d, key)
    if v is None:
        return default
    s = str(v).strip().lower()
    return "Yes" if s in ("yes", "1", "true", "y") else "No"


class CRMImporter:
    """
    Loads import.xlsx and parses each sheet into typed Pydantic records.
    Validation errors are collected (non-fatal); valid records are returned.
    """

    SHEET_COMPANIES = "Companies"
    SHEET_CONTACTS = "Contacts"
    SHEET_GIGS_LEADS = "Gigs_Leads"
    SHEET_CALLS = "Calls"
    SHEET_NOTES = "Notes"

    def __init__(self, path: str | Path):
        self.path = Path(path)

    def load_all(self) -> ImportResult:
        if not self.path.exists():
            raise FileNotFoundError(f"CRM file not found: {self.path}")
        wb = openpyxl.load_workbook(self.path, data_only=True)
        result = ImportResult()

        if self.SHEET_COMPANIES in wb.sheetnames:
            rows = _sheet_to_dicts(wb[self.SHEET_COMPANIES])
            for i, row in enumerate(rows):
                try:
                    result.venues.append(self._parse_venue(row))
                except (ValidationError, Exception) as e:
                    result.errors.append({"sheet": "Companies", "row": i + 2, "error": str(e)})

        if self.SHEET_CONTACTS in wb.sheetnames:
            rows = _sheet_to_dicts(wb[self.SHEET_CONTACTS])
            for i, row in enumerate(rows):
                try:
                    result.contacts.append(self._parse_contact(row))
                except (ValidationError, Exception) as e:
                    result.errors.append({"sheet": "Contacts", "row": i + 2, "error": str(e)})

        if self.SHEET_GIGS_LEADS in wb.sheetnames:
            rows = _sheet_to_dicts(wb[self.SHEET_GIGS_LEADS])
            for i, row in enumerate(rows):
                try:
                    result.gig_leads.append(self._parse_gig_lead(row))
                except (ValidationError, Exception) as e:
                    result.errors.append({"sheet": "Gigs_Leads", "row": i + 2, "error": str(e)})

        if self.SHEET_CALLS in wb.sheetnames:
            rows = _sheet_to_dicts(wb[self.SHEET_CALLS])
            for i, row in enumerate(rows):
                try:
                    result.calls.append(self._parse_call(row))
                except (ValidationError, Exception) as e:
                    result.errors.append({"sheet": "Calls", "row": i + 2, "error": str(e)})

        if self.SHEET_NOTES in wb.sheetnames:
            rows = _sheet_to_dicts(wb[self.SHEET_NOTES])
            for i, row in enumerate(rows):
                try:
                    result.notes.append(self._parse_note(row))
                except (ValidationError, Exception) as e:
                    result.errors.append({"sheet": "Notes", "row": i + 2, "error": str(e)})

        logger.info(
            "Import complete: %d records loaded, %d errors",
            result.total_records,
            result.error_count,
        )
        return result

    # ------------------------------------------------------------------
    def _parse_venue(self, row: dict) -> VenueRecord:
        return VenueRecord(
            lead_id=_int(row, "UniqueId") or 0,
            company=_str(row, "Name") or "(unnamed)",
            website=_str(row, "Website"),
            segment=_str(row, "Record Type"),
            phone=_str(row, "Phone"),
            city=_str(row, "Billing City"),
            state=_str(row, "Billing State"),
            notes=_str(row, "Notes"),
            seating_capacity_tier=_str(row, "Custom1"),
            mission_excerpt=_str(row, "Custom2"),
            programming_focus=_str(row, "Custom3"),
            venue_name=_str(row, "Custom4"),
        )

    def _parse_contact(self, row: dict) -> ContactRecord:
        return ContactRecord(
            contact_id=_int(row, "UniqueId") or 0,
            first_name=_str(row, "First Name"),
            last_name=_str(row, "Last Name"),
            company_id=_int(row, "Company UniqueId"),
            source=_str(row, "Lead Source"),
            phone=_str(row, "Work Phone"),
            email=_str(row, "Email"),
            email_opt_out=_yes_no(row, "Email Opt Out"),
            title=_str(row, "Title"),
            sequence_id=_str(row, "Custom1"),
            step_number=_int(row, "Custom2"),
            email_open=_yes_no(row, "Custom3"),
            email_click=_yes_no(row, "Custom4"),
            email_reply=_yes_no(row, "Custom5"),
            reply_category=_str(row, "Custom6"),
            last_updated=_str(row, "Custom7"),
            audit_trail_id=_str(row, "Custom8"),
        )

    def _parse_gig_lead(self, row: dict) -> GigLeadRecord:
        show_date = row.get("Show Date")
        if isinstance(show_date, str):
            try:
                show_date = datetime.fromisoformat(show_date)
            except ValueError:
                show_date = None

        return GigLeadRecord(
            gig_lead_id=_int(row, "UniqueId") or 0,
            name=_str(row, "Name"),
            status=_str(row, "Status"),
            company_id=_int(row, "Company UniqueId"),
            contact_id=_int(row, "Primary Contact UniqueId"),
            source=_str(row, "Lead Source"),
            show_date=show_date,
            venue_name=_str(row, "Venue Name"),
            final_fee=_float(row, "Fee"),
            projected_fee=_float(row, "Quote1"),
            deposit=_float(row, "Deposit"),
            contract_sent=_str(row, "Contract sent"),
            contract_received=_str(row, "Contract Received"),
            outcome=_str(row, "Custom1"),
            meeting_booked=_yes_no(row, "Custom2"),
            show_booked=_yes_no(row, "Custom3"),
            send_time=_str(row, "Custom4"),
        )

    def _parse_call(self, row: dict) -> CallRecord:
        return CallRecord(
            call_id=_int(row, "UniqueId") or 0,
            subject=_str(row, "Subject"),
            call_datetime=row.get("CallDateTime") if isinstance(row.get("CallDateTime"), datetime) else None,
            status=_str(row, "Call Status"),
            notes=_str(row, "Call Notes"),
            gig_id=_int(row, "Related Gig"),
            contact_id=_int(row, "Related Contact"),
            company_id=_int(row, "Related Company"),
        )

    def _parse_note(self, row: dict) -> NoteRecord:
        return NoteRecord(
            note_id=_int(row, "UniqueId") or 0,
            subject=_str(row, "Subject"),
            note_date=row.get("Date").date() if hasattr(row.get("Date"), "date") else None,
            notes=_str(row, "Notes"),
            gig_id=_int(row, "Related Gig"),
            contact_id=_int(row, "Related Contact"),
            company_id=_int(row, "Related Company"),
        )
